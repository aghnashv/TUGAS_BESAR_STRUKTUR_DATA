# music_app_final.py
import sys
import os
import random
from dataclasses import dataclass
from typing import Optional, Dict, List
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# -----------------------
# Config / Paths
# -----------------------
DB_DIR = "database"
SONGS_FILE = os.path.join(DB_DIR, "songs.xlsx")
PLAYLISTS_FILE = os.path.join(DB_DIR, "playlists.xlsx")
PLAYLIST_SONGS_FILE = os.path.join(DB_DIR, "playlist_songs.xlsx")

# -----------------------
# Util
# -----------------------
def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")

# -----------------------
# Data classes
# -----------------------
@dataclass
class Song:
    song_id: str
    title: str
    artist: str
    genre: str
    album: str
    year: int

# -----------------------
# Doubly Linked List (used for library order & playlists)
# -----------------------
class DLLNode:
    def __init__(self, data: Song):
        self.data: Song = data
        self.prev: Optional["DLLNode"] = None
        self.next: Optional["DLLNode"] = None

class DoublyLinkedList:
    def __init__(self):
        self.head: Optional[DLLNode] = None
        self.tail: Optional[DLLNode] = None
        self.size = 0

    def append(self, node: DLLNode):
        if self.tail is None:
            self.head = self.tail = node
        else:
            self.tail.next = node
            node.prev = self.tail
            self.tail = node
        self.size += 1

    def remove(self, node: DLLNode):
        if node.prev:
            node.prev.next = node.next
        else:
            self.head = node.next

        if node.next:
            node.next.prev = node.prev
        else:
            self.tail = node.prev

        node.prev = node.next = None
        self.size -= 1

    def to_list(self) -> List[Song]:
        res = []
        cur = self.head
        while cur:
            res.append(cur.data)
            cur = cur.next
        return res

# -----------------------
# Stack (baru) - untuk history pemutaran
# -----------------------
class Stack:
    def __init__(self):
        self.items = []

    def push(self, item):
        self.items.append(item)

    def pop(self):
        if not self.is_empty():
            return self.items.pop()
        return None

    def peek(self):
        if not self.is_empty():
            return self.items[-1]
        return None

    def is_empty(self):
        return len(self.items) == 0

    def __len__(self):
        return len(self.items)

# -----------------------
# BST for title -> list of song_ids
# -----------------------
class BSTNode:
    def __init__(self, key: str, song_ids: Optional[List[str]] = None):
        self.key = key
        self.song_ids = song_ids or []
        self.left: Optional["BSTNode"] = None
        self.right: Optional["BSTNode"] = None

class SongBST:
    def __init__(self):
        self.root: Optional[BSTNode] = None

    def insert(self, title_key: str, song_id: str):
        self.root = self._insert(self.root, title_key, song_id)

    def _insert(self, node: Optional[BSTNode], key: str, song_id: str) -> BSTNode:
        if node is None:
            return BSTNode(key, [song_id])
        if key < node.key:
            node.left = self._insert(node.left, key, song_id)
        elif key > node.key:
            node.right = self._insert(node.right, key, song_id)
        else:
            if song_id not in node.song_ids:
                node.song_ids.append(song_id)
        return node

    def search(self, key: str) -> Optional[List[str]]:
        node = self.root
        while node:
            if key < node.key:
                node = node.left
            elif key > node.key:
                node = node.right
            else:
                return node.song_ids
        return None

    def delete(self, key: str, song_id: str):
        self.root = self._delete(self.root, key, song_id)

    def _delete(self, node: Optional[BSTNode], key: str, song_id: str) -> Optional[BSTNode]:
        if node is None:
            return None
        if key < node.key:
            node.left = self._delete(node.left, key, song_id)
        elif key > node.key:
            node.right = self._delete(node.right, key, song_id)
        else:
            # remove song_id
            if song_id in node.song_ids:
                node.song_ids.remove(song_id)
            # if empty, remove node from tree
            if len(node.song_ids) == 0:
                # two children
                if node.left and node.right:
                    succ = self._min_node(node.right)
                    node.key = succ.key
                    node.song_ids = succ.song_ids.copy()
                    # delete one of successor's song ids via recursive call
                    node.right = self._delete(node.right, succ.key, succ.song_ids[0])
                else:
                    return node.left or node.right
        return node

    def _min_node(self, node: BSTNode) -> BSTNode:
        current = node
        while current.left:
            current = current.left
        return current

# -----------------------
# Library (with DLL + map + BST)
# -----------------------
class Library:
    def __init__(self):
        self.dll = DoublyLinkedList()
        self.map: Dict[str, DLLNode] = {}   # song_id -> DLLNode
        self.bst = SongBST()

    def add_song(self, song: Song) -> bool:
        if song.song_id in self.map:
            return False
        node = DLLNode(song)
        self.dll.append(node)
        self.map[song.song_id] = node
        # insert title to BST
        self.bst.insert(song.title.lower(), song.song_id)
        return True

    def view_songs(self) -> List[Song]:
        return self.dll.to_list()

    def get_song(self, song_id: str) -> Optional[Song]:
        node = self.map.get(song_id)
        return node.data if node else None

    def update_song(self, song_id: str, **kwargs) -> bool:
        node = self.map.get(song_id)
        if not node:
            return False
        old_title = node.data.title
        # update fields
        for k, v in kwargs.items():
            if v is not None and hasattr(node.data, k):
                setattr(node.data, k, v)
        # if title changed, update BST
        if 'title' in kwargs and kwargs['title'] and kwargs['title'] != old_title:
            self.bst.delete(old_title.lower(), song_id)
            self.bst.insert(kwargs['title'].lower(), song_id)
        return True

    def delete_song(self, song_id: str) -> bool:
        node = self.map.get(song_id)
        if not node:
            return False
        # delete from BST
        self.bst.delete(node.data.title.lower(), song_id)
        # remove from dll and map
        self.dll.remove(node)
        del self.map[song_id]
        return True

    def search_by_title_bst(self, title: str) -> Optional[List[str]]:
        return self.bst.search(title.lower())

    def find_similar(self, reference: Song) -> Optional[Song]:
        # same logic as before
        if self.dll.size <= 1:
            return None
        cur = self.dll.head
        artist_matches = []
        genre_matches = []
        others = []
        while cur:
            s = cur.data
            if s.song_id == reference.song_id:
                cur = cur.next
                continue
            if s.artist.lower() == reference.artist.lower():
                artist_matches.append(s)
            elif s.genre.lower() == reference.genre.lower():
                genre_matches.append(s)
            else:
                others.append(s)
            cur = cur.next
        if artist_matches:
            return random.choice(artist_matches)
        if genre_matches:
            return random.choice(genre_matches)
        if others:
            return random.choice(others)
        return None

# -----------------------
# Playlist
# -----------------------
class Playlist:
    def __init__(self, name: str):
        self.name = name
        self.dll = DoublyLinkedList()
        self.map: Dict[str, DLLNode] = {}

    def add_song_node(self, library_node: DLLNode) -> bool:
        sid = library_node.data.song_id
        if sid in self.map:
            return False
        node = DLLNode(library_node.data)
        self.dll.append(node)
        self.map[sid] = node
        return True

    def remove_song(self, song_id: str) -> bool:
        node = self.map.get(song_id)
        if not node:
            return False
        self.dll.remove(node)
        del self.map[song_id]
        return True

    def list_songs(self) -> List[Song]:
        return self.dll.to_list()

    def cleanup_removed_song(self, song_id: str):
        if song_id in self.map:
            node = self.map[song_id]
            self.dll.remove(node)
            del self.map[song_id]

# -----------------------
# Player (menggunakan Stack untuk history)
# -----------------------
class Player:
    def __init__(self, library: Library):
        self.library = library
        self.current_node: Optional[DLLNode] = None  # when playing from library
        self.is_playing: bool = False
        self.current_playlist: Optional[Playlist] = None
        self.playlist_cursor: Optional[DLLNode] = None

        # mengganti List[str] dengan Stack
        self.history = Stack()

    def play_from_library(self, song_id: str):
        node = self.library.map.get(song_id)
        if not node:
            print("Lagu tidak ditemukan di library.")
            return
        self.current_playlist = None
        self.playlist_cursor = None
        self.current_node = node
        self.is_playing = True

        # PUSH ke stack
        self.history.push(song_id)

        print(f"▶ Memutar: {node.data.title} — {node.data.artist}")

    def play_from_playlist(self, playlist: Playlist, song_id: Optional[str] = None):
        if playlist.dll.size == 0:
            print("Playlist kosong.")
            return
        self.current_playlist = playlist
        if song_id:
            node = playlist.map.get(song_id)
            if not node:
                print("Lagu tidak ada di playlist.")
                return
            self.playlist_cursor = node
        else:
            self.playlist_cursor = playlist.dll.head

        self.current_node = self.playlist_cursor
        self.is_playing = True

        # PUSH ke stack
        self.history.push(self.current_node.data.song_id)

        print(f"▶ Memutar dari playlist '{playlist.name}': {self.current_node.data.title} — {self.current_node.data.artist}")

    def stop(self):
        if not self.is_playing:
            print("Tidak sedang memutar lagu.")
            return
        print(f"⏹ Menghentikan: {self.get_current_info()}")
        self.is_playing = False

    def next(self):
        # jika sedang di playlist
        if self.current_playlist and self.playlist_cursor:
            if self.playlist_cursor.next:
                self.playlist_cursor = self.playlist_cursor.next
                self.current_node = self.playlist_cursor
                self.is_playing = True

                # PUSH next track
                self.history.push(self.current_node.data.song_id)

                print(f"▶ Next (playlist): {self.get_current_info()}")
            else:
                print("Sudah di lagu terakhir playlist.")
            return

        # jika tidak di playlist → cari similar
        if self.current_node:
    # NEXT berurutan di library
            if self.current_node.next:
                self.current_node = self.current_node.next
                self.is_playing = True

                # PUSH next track ke history
                self.history.push(self.current_node.data.song_id)

                print(f"▶ Next: {self.get_current_info()}")
            else:
                print("Sudah di lagu terakhir library.")
        else:
            print("Tidak ada lagu yang sedang dipilih.")

    def prev(self):
        # minimal ada 2 lagu di stack
        if len(self.history) <= 1:
            print("Tidak ada lagu sebelumnya.")
            return

        # POP lagu sekarang
        self.history.pop()

        # Peek lagu sebelumnya
        prev_id = self.history.peek()
        node = self.library.map.get(prev_id)

        if node:
            self.current_node = node
            self.is_playing = True
            print(f"▶ Prev: {self.get_current_info()}")
        else:
            print("Lagu sebelumnya tidak tersedia lagi.")

    def get_current_info(self) -> str:
        if not self.current_node:
            return "<tidak ada>"
        s = self.current_node.data
        return f"{s.title} — {s.artist} ({s.album}, {s.year})"

    def cleanup_if_current_deleted(self, song_id: str):
        if self.current_node and self.current_node.data.song_id == song_id:
            print("Lagu yang sedang diputar dihapus. Memutar dihentikan.")
            self.stop()
            self.current_node = None
            self.playlist_cursor = None
            self.current_playlist = None
            self.history = Stack()  # reset stack

# -----------------------
# Application CLI (integrated)
# -----------------------
class MusicAppCLI:
    def __init__(self):
        os.makedirs(DB_DIR, exist_ok=True)
        self.library = Library()
        self.playlists: Dict[str, Playlist] = {}
        self.player = Player(self.library)
        self.admin_password = "admin123"
        # load data from excel (if any)
        self.load_data_from_excel()
        # if library empty after load, seed demo then save
        if not self.library.view_songs():
            self._seed_demo()
            self.save_all_to_excel()

    def _seed_demo(self):
        demos = [
            Song("S001", "Bintang Kecil", "Penyanyi A", "Pop", "Album A", 2020),
            Song("S002", "Langit Biru", "Penyanyi B", "Indie", "Album B", 2021),
            Song("S003", "Rindu", "Penyanyi A", "Pop", "Album A", 2019),
            Song("S004", "Rindu", "Penyanyi B", "Pop", "Album X", 2018),
            Song("S005", "Malam Sunyi", "Penyanyi C", "Jazz", "Album C", 2017),
        ]
        for s in demos:
            self.library.add_song(s)

    # -----------------------
    # Excel helpers
    # -----------------------
    def save_songs_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["song_id", "title", "artist", "genre", "album", "year"])
        for s in self.library.view_songs():
            ws.append([s.song_id, s.title, s.artist, s.genre, s.album, s.year])
        wb.save(SONGS_FILE)

    def save_playlists_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["playlist_name"])
        for name in self.playlists.keys():
            ws.append([name])
        wb.save(PLAYLISTS_FILE)

    def save_playlist_songs_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["playlist_name", "song_id"])
        for name, pl in self.playlists.items():
            for s in pl.list_songs():
                ws.append([name, s.song_id])
        wb.save(PLAYLIST_SONGS_FILE)

    def save_all_to_excel(self):
        try:
            self.save_songs_to_excel()
            self.save_playlists_to_excel()
            self.save_playlist_songs_to_excel()
        except Exception as e:
            print("Gagal menyimpan data:", e)

    def load_data_from_excel(self):
        # load songs (also inserts into BST via library.add_song)
        try:
            if os.path.exists(SONGS_FILE):
                wb = load_workbook(SONGS_FILE)
                ws = wb.active
                first = True
                for row in ws.iter_rows(values_only=True):
                    if first:
                        first = False
                        continue
                    if not row or row[0] is None:
                        continue
                    sid, title, artist, genre, album, year = row
                    try:
                        year_int = int(year)
                    except:
                        year_int = 0
                    # avoid duplicates when loading multiple times
                    if str(sid) not in self.library.map:
                        self.library.add_song(Song(str(sid), str(title), str(artist), str(genre), str(album), year_int))
        except Exception:
            pass

        # load playlists
        try:
            if os.path.exists(PLAYLISTS_FILE):
                wb = load_workbook(PLAYLISTS_FILE)
                ws = wb.active
                first = True
                for row in ws.iter_rows(values_only=True):
                    if first:
                        first = False
                        continue
                    if not row or row[0] is None:
                        continue
                    pl_name = str(row[0])
                    if pl_name not in self.playlists:
                        self.playlists[pl_name] = Playlist(pl_name)
        except Exception:
            pass

        # load playlist songs
        try:
            if os.path.exists(PLAYLIST_SONGS_FILE):
                wb = load_workbook(PLAYLIST_SONGS_FILE)
                ws = wb.active
                first = True
                for row in ws.iter_rows(values_only=True):
                    if first:
                        first = False
                        continue
                    if not row or row[0] is None or row[1] is None:
                        continue
                    pl_name = str(row[0])
                    sid = str(row[1])
                    if pl_name not in self.playlists:
                        self.playlists[pl_name] = Playlist(pl_name)
                    if sid in self.library.map:
                        self.playlists[pl_name].add_song_node(self.library.map[sid])
        except Exception:
            pass

    # -----------------------
    # MAIN / Menus
    # -----------------------
    def run(self):
        while True:
            clear_screen()
            print("=== MENU UTAMA ===")
            print("1. Login Admin")
            print("2. Login User")
            print("0. Keluar")
            ch = input("Pilih: ").strip()
            if ch == "1":
                self.admin_page()
            elif ch == "2":
                self.user_page()
            elif ch == "0":
                print("Keluar... Terima Kasih!!")
                self.save_all_to_excel()
                sys.exit(0)
            else:
                print("Pilihan tidak valid.")
                input("\nENTER...")

    # -----------------------
    # Admin
    # -----------------------
    def admin_page(self):
        clear_screen()
        print("Tekan 0 untuk kembali.")
        pw = input("Masukkan password admin: ").strip()
        if pw == "0":
            return
        if pw != self.admin_password:
            print("Password salah! Kembali.")
            input("\nENTER...")
            return

        while True:
            clear_screen()
            print("--- ADMIN ---")
            print("1. Tambah Lagu")
            print("2. Lihat Semua Lagu")
            print("3. Ubah Data Lagu")
            print("4. Hapus Lagu")
            print("0. Kembali")
            c = input("Pilih: ").strip()
            if c == "1":
                self._admin_add_song()
            elif c == "2":
                self._admin_view_songs()
                input("\nENTER...")
            elif c == "3":
                self._admin_update_song()
            elif c == "4":
                self._admin_delete_song()
            elif c == "0":
                return
            else:
                print("Pilihan tidak valid.")
                input("\nENTER...")

    def _admin_add_song(self):
        clear_screen()
        print("Tekan 0 untuk kembali.")
        sid = input("ID lagu: ").strip()
        if sid == "0": return
        title = input("Judul: ").strip()
        if title == "0": return
        artist = input("Penyanyi: ").strip()
        if artist == "0": return
        genre = input("Genre: ").strip()
        if genre == "0": return
        album = input("Album: ").strip()
        if album == "0": return
        year_in = input("Tahun rilis: ").strip()
        if year_in == "0": return
        try:
            year = int(year_in)
        except:
            print("Tahun harus angka.")
            input("\nENTER...")
            return
        ok = self.library.add_song(Song(sid, title, artist, genre, album, year))
        if ok:
            print("Lagu berhasil ditambahkan.")
            self.save_songs_to_excel()
        else:
            print("ID sudah ada.")
        input("\nENTER...")

    def _admin_view_songs(self):
        clear_screen()
        print("=== DAFTAR LAGU ===")
        songs = self.library.view_songs()
        if not songs:
            print("Library kosong.")
            return
        for s in songs:
            print(f"- {s.song_id}: {s.title} — {s.artist} ({s.genre}, {s.album}, {s.year})")

    def _admin_update_song(self):
        clear_screen()
        print("Tekan 0 untuk kembali.")
        sid = input("Masukkan ID lagu yang ingin diubah: ").strip()
        if sid == "0": return
        node = self.library.map.get(sid)
        if not node:
            print("Lagu tidak ditemukan.")
            input("\nENTER...")
            return
        s = node.data
        print("Biarkan kosong jika tidak ingin mengubah field tersebut.")
        title = input(f"Judul ({s.title}): ").strip() or None
        artist = input(f"Penyanyi ({s.artist}): ").strip() or None
        genre = input(f"Genre ({s.genre}): ").strip() or None
        album = input(f"Album ({s.album}): ").strip() or None
        year_str = input(f"Tahun ({s.year}): ").strip() or None
        updates = {}
        if title is not None: updates['title'] = title
        if artist is not None: updates['artist'] = artist
        if genre is not None: updates['genre'] = genre
        if album is not None: updates['album'] = album
        if year_str:
            try:
                updates['year'] = int(year_str)
            except:
                print("Tahun harus angka. Perubahan tahun diabaikan.")
        ok = self.library.update_song(sid, **updates)
        if ok:
            print("Data lagu berhasil diperbarui.")
            self.save_songs_to_excel()
            # saving playlist_songs is harmless (meta changed)
            self.save_playlist_songs_to_excel()
        else:
            print("Gagal memperbarui lagu.")
        input("\nENTER...")

    def _admin_delete_song(self):
        clear_screen()
        print("Tekan 0 untuk kembali.")
        sid = input("Masukkan ID lagu yang ingin dihapus: ").strip()
        if sid == "0": return
        if sid not in self.library.map:
            print("Lagu tidak ditemukan.")
            input("\nENTER...")
            return
        ok = self.library.delete_song(sid)
        if ok:
            # remove from playlists
            for pl in self.playlists.values():
                pl.cleanup_removed_song(sid)
            # if currently playing, stop
            self.player.cleanup_if_current_deleted(sid)
            # save
            self.save_songs_to_excel()
            self.save_playlist_songs_to_excel()
            print("Lagu berhasil dihapus dari library dan semua playlist.")
        else:
            print("Gagal menghapus lagu.")
        input("\nENTER...")

    # -----------------------
    # USER PAGE
    # -----------------------
    def user_page(self):
        while True:
            clear_screen()
            print("--- USER PAGE ---")
            print("1. Buat playlist")
            print("2. Tambah lagu ke playlist")
            print("3. Hapus lagu dari playlist")
            print("4. Lihat playlist")
            print("5. Cari")
            print("6. Play lagu (dari library)")
            print("7. Play dari playlist")
            print("8. Hapus playlist")
            print("0. Kembali")
            ch = input("Pilih: ").strip()
            if ch == "1":
                self._user_create_playlist()
            elif ch == "2":
                self._user_add_song_to_playlist()
            elif ch == "3":
                self._user_remove_song_from_playlist()
            elif ch == "4":
                self._user_view_playlists()
                input("\nENTER...")
            elif ch == "5":
                self._user_search()   # using BST
                input("\nENTER...")
            elif ch == "6":
                self._user_play_from_library()
            elif ch == "7":
                self._user_play_from_playlist()
            elif ch == "11":
                self._user_delete_playlist()
            elif ch == "0":
                return
            else:
                print("Pilihan tidak valid.")
                input("\nENTER...")

    def _user_create_playlist(self):
        clear_screen()
        print("Tekan 0 untuk kembali.")
        name = input("Nama playlist: ").strip()
        if name == "0": return
        if not name:
            print("Nama tidak boleh kosong.")
            input("\nENTER...")
            return
        if name in self.playlists:
            print("Nama playlist sudah ada.")
            input("\nENTER...")
            return
        self.playlists[name] = Playlist(name)
        self.save_playlists_to_excel()
        print(f"Playlist '{name}' dibuat.")
        input("\nENTER...")

    def _user_add_song_to_playlist(self):
        clear_screen()
        if not self.library.map:
            print("Library kosong. Tidak ada lagu untuk ditambahkan.")
            input("\nENTER...")
            return
        if not self.playlists:
            print("Anda belum memiliki playlist. Buat playlist terlebih dahulu.")
            input("\nENTER...")
            return
        # show playlists
        print("Daftar Playlist:")
        for name in self.playlists:
            print(f"- {name}")
        print("\nTekan 0 untuk kembali.")
        name = input("Pilih playlist: ").strip()
        if name == "0": return
        pl = self.playlists.get(name)
        if not pl:
            print("Playlist tidak ditemukan.")
            input("\nENTER...")
            return
        print("\nTambah berdasarkan (1) ID atau (2) Judul")
        print("Tekan 0 untuk kembali.")
        method = input("Pilih: ").strip()
        if method == "0": return
        node = None
        if method == "1":
            sid = input("ID lagu: ").strip()
            if sid == "0": return
            node = self.library.map.get(sid)
            if not node:
                print("Lagu tidak ditemukan.")
                input("\nENTER...")
                return
        else:
            keyword = input("Judul (substring): ").strip().lower()
            if keyword == "0": return
            candidates = [s for s in self.library.view_songs() if keyword in s.title.lower()]
            if not candidates:
                print("Tidak ada lagu ditemukan.")
                input("\nENTER...")
                return
            print("\nHasil pencarian:")
            for i, s in enumerate(candidates, 1):
                print(f"{i}. {s.song_id}: {s.title} — {s.artist}")
            try:
                idx = int(input("Pilih nomor (0 untuk batal): ").strip())
                if idx == 0: return
                sid = candidates[idx - 1].song_id
            except:
                print("Input salah.")
                input("\nENTER...")
                return
            node = self.library.map.get(sid)
        ok = pl.add_song_node(node)
        if ok:
            print("Lagu ditambahkan ke playlist.")
            self.save_playlist_songs_to_excel()
        else:
            print("Lagu sudah ada di playlist.")
        input("\nENTER...")

    def _user_remove_song_from_playlist(self):
        clear_screen()
        print("Tekan 0 untuk kembali.")
        name = input("Nama playlist: ").strip()
        if name == "0": return
        pl = self.playlists.get(name)
        if not pl:
            print("Playlist tidak ditemukan.")
            input("\nENTER...")
            return
        sid = input("ID lagu yang ingin dihapus dari playlist: ").strip()
        if sid == "0": return
        ok = pl.remove_song(sid)
        if ok:
            print("Lagu dihapus dari playlist.")
            self.save_playlist_songs_to_excel()
            # if currently playing that playlist and that song, move or stop
            if self.player.current_playlist == pl and self.player.playlist_cursor and self.player.playlist_cursor.data.song_id == sid:
                print("Lagu yang sedang diputar dihapus dari playlist. Memilih lagu berikutnya jika ada...")
                if self.player.playlist_cursor.next:
                    self.player.playlist_cursor = self.player.playlist_cursor.next
                    self.player.current_node = self.player.playlist_cursor
                    print(f"Sekarang memutar: {self.player.get_current_info()}")
                else:
                    self.player.stop()
                    self.player.current_node = None
        else:
            print("Lagu tidak ada di playlist.")
        input("\nENTER...")

    def _user_view_playlists(self):
        clear_screen()
        if not self.playlists:
            print("Belum ada playlist.")
            return
        for name, pl in self.playlists.items():
            print(f"\nPlaylist: {name} (jumlah: {pl.dll.size})")
            for s in pl.list_songs():
                print(f"- {s.song_id}: {s.title} — {s.artist}")

    def _user_delete_playlist(self):
        clear_screen()
        print("Tekan 0 untuk kembali.")
        name = input("Nama playlist yang akan dihapus: ").strip()
        if name == "0": return
        if name not in self.playlists:
            print("Playlist tidak ditemukan.")
            input("\nENTER...")
            return
        if self.player.current_playlist and self.player.current_playlist.name == name:
            self.player.stop()
            self.player.current_node = None
            self.player.playlist_cursor = None
            self.player.current_playlist = None
            self.player.history = Stack()
        del self.playlists[name]
        self.save_playlists_to_excel()
        self.save_playlist_songs_to_excel()
        print("Playlist dihapus.")
        input("\nENTER...")

    def _user_search(self):
        clear_screen()
        print("=== MENU CARI ===")
        print("1. Cari berdasarkan artis")
        print("2. Cari berdasarkan judul")
        print("0. Kembali")
        opsi = input("Pilih: ").strip()

        if opsi == "0":
            return

    # ============================
    # 1. CARI BERDASARKAN ARTIS
    # ============================
        if opsi == "1":
            clear_screen()
            print("=== CARI BERDASARKAN ARTIS ===")
            artist = input("Nama artis/penyanyi : ").strip().lower()

            if artist == "0":
                return

            hasil = [
                s for s in self.library.view_songs()
                if artist in s.artist.lower()
            ]

            clear_screen()
            print("=== HASIL CARI ARTIS ===")
            if not hasil:
                print("Tidak ada lagu dari artis tersebut.")
                return

            for s in hasil:
                print(f"- {s.song_id}: {s.title} — {s.artist} ({s.album}, {s.year})")
            return

    # ============================
    # 2. CARI BERDASARKAN JUDUL
    # ============================
        if opsi == "2":
            clear_screen()
            print("=== CARI BERDASARKAN JUDUL ===")
            key = input("Masukkan judul : ").strip().lower()
            if key == "0":
                return

        # 1) coba exact BST
            song_ids = self.library.search_by_title_bst(key)

            clear_screen()
            print("=== HASIL CARI JUDUL ===")
            if song_ids:
                for sid in song_ids:
                    s = self.library.get_song(sid)
                    if s:
                        print(f"- {s.song_id}: {s.title} — {s.artist} ({s.album}, {s.year})")
                return

        # 2) fallback substring
            hasil = [
                s for s in self.library.view_songs()
                if key in s.title.lower()
            ]

            if hasil:
                print("(Hasil substring):")
                for s in hasil:
                    print(f"- {s.song_id}: {s.title} — {s.artist} ({s.album}, {s.year})")
            else:
                print("Tidak ada lagu dengan judul tersebut.")
            return

        print("Pilihan tidak valid.")
        input("\nENTER...")

    def _user_play_from_library(self):
        clear_screen()
        print("Tekan 0 untuk kembali.")
        title = input("Masukkan judul lagu (substring atau judul lengkap): ").strip().lower()
        if title == "0": return
        candidates = [s for s in self.library.view_songs() if title in s.title.lower()]
        if not candidates:
            print("Tidak ada lagu dengan judul tersebut.")
            input("\nENTER...")
            return
        if len(candidates) == 1:
            self.player.play_from_library(candidates[0].song_id)
            self._player_controls()
            return
        # multiple candidates
        print("\nDitemukan beberapa lagu:")
        for i, s in enumerate(candidates, 1):
            print(f"{i}. {s.song_id}: {s.title} — {s.artist}")
        try:
            idx = int(input("Pilih nomor (0 batal): ").strip())
            if idx == 0:
                return
            if 1 <= idx <= len(candidates):
                self.player.play_from_library(candidates[idx-1].song_id)
                self._player_controls()
            else:
                print("Pilihan tidak valid.")
                input("\nENTER...")
        except:
            print("Input tidak valid.")
            input("\nENTER...")

    def _user_play_from_playlist(self):
        clear_screen()
        if not self.playlists:
            print("Belum ada playlist.")
            input("\nENTER...")
            return
        print("Daftar playlist:")
        for name in self.playlists:
            print(f"- {name}")
        name = input("\nNama playlist: ").strip()
        if name == "0": return
        pl = self.playlists.get(name)
        if not pl:
            print("Playlist tidak ditemukan.")
            input("\nENTER...")
            return
        sid = input("ID lagu di playlist (kosong -> mulai dari awal): ").strip()
        if sid == "": sid = None
        self.player.play_from_playlist(pl, sid)
        self._player_controls()

    # -----------------------
    # PLAYER CONTROLS (sub-menu saat memutar)
    # -----------------------
    def _player_controls(self):
        while True:
            clear_screen()
            print("\n=== PLAYER CONTROL ===")
            print("Sekarang:", self.player.get_current_info())
            print("1. Stop")
            print("2. Next")
            print("3. Prev")
            print("0. Kembali")
            c = input("Pilih: ").strip()

            if c == "1":
                self.player.stop()
            elif c == "2":
                self.player.next()
            elif c == "3":
                self.player.prev()
            elif c == "0":
                return
            else:
                print("Pilihan tidak valid.")

    # -----------------------
    # Admin/User actions that change data -> call saves
    # -----------------------
    # helper wrappers to ensure saving consistently
    def save_songs_to_excel(self):
        self.save_songs_to_excel = self.save_songs_to_excel_impl
        self.save_songs_to_excel_impl()

    def save_playlists_to_excel(self):
        self.save_playlists_to_excel = self.save_playlists_to_excel_impl
        self.save_playlists_to_excel_impl()

    def save_playlist_songs_to_excel(self):
        self.save_playlist_songs_to_excel = self.save_playlist_songs_to_excel_impl
        self.save_playlist_songs_to_excel_impl()

    # implementations (kept separate to avoid recursive name confusion)
    def save_songs_to_excel_impl(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["song_id", "title", "artist", "genre", "album", "year"])
        for s in self.library.view_songs():
            ws.append([s.song_id, s.title, s.artist, s.genre, s.album, s.year])
        wb.save(SONGS_FILE)

    def save_playlists_to_excel_impl(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["playlist_name"])
        for name in self.playlists.keys():
            ws.append([name])
        wb.save(PLAYLISTS_FILE)

    def save_playlist_songs_to_excel_impl(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["playlist_name", "song_id"])
        for name, pl in self.playlists.items():
            for s in pl.list_songs():
                ws.append([name, s.song_id])
        wb.save(PLAYLIST_SONGS_FILE)

    # keep old save_all_to_excel name
    def save_all_to_excel(self):
        try:
            self.save_songs_to_excel_impl()
            self.save_playlists_to_excel_impl()
            self.save_playlist_songs_to_excel_impl()
        except Exception as e:
            print("Gagal menyimpan:", e)

# -----------------------
# run
# -----------------------
if __name__ == "__main__":
    app = MusicAppCLI()
    app.run()