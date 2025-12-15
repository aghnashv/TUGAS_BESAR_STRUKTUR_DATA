import sys
import os
from dataclasses import dataclass
from typing import Optional, Dict, List
from openpyxl import Workbook, load_workbook
import pygame

# =======================
# PATH CONFIG
# =======================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SONG_DIR = os.path.join(BASE_DIR, "songs")
DB_DIR = os.path.join(BASE_DIR, "database")
SONGS_FILE = os.path.join(DB_DIR, "songs.xlsx")

os.makedirs(DB_DIR, exist_ok=True)
os.makedirs(SONG_DIR, exist_ok=True)

# =======================
# HELPER
# =======================
def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")

# =======================
# DATA CLASS
# =======================
@dataclass
class Song:
    song_id: str
    title: str
    artist: str
    genre: str
    album: str
    year: int

# =======================
# DOUBLY LINKED LIST
# =======================
class DLLNode:
    def __init__(self, data: Song):
        self.data = data
        self.prev = None
        self.next = None

class DoublyLinkedList:
    def __init__(self):
        self.head = None
        self.tail = None

    def append(self, node: DLLNode):
        if not self.head:
            self.head = self.tail = node
        else:
            self.tail.next = node
            node.prev = self.tail
            self.tail = node

    def remove(self, node: DLLNode):
        if node.prev:
            node.prev.next = node.next
        else:
            self.head = node.next
        if node.next:
            node.next.prev = node.prev
        else:
            self.tail = node.prev

    def to_list(self):
        cur = self.head
        res = []
        while cur:
            res.append(cur.data)
            cur = cur.next
        return res

# =======================
# STACK (PREV SONG)
# =======================
class Stack:
    def __init__(self):
        self.items = []
    def push(self, x):
        self.items.append(x)
    def pop(self):
        return self.items.pop() if self.items else None
    def peek(self):
        return self.items[-1] if self.items else None
    def __len__(self):
        return len(self.items)
    def is_empty(self):
        return len(self.items) == 0

# =======================
# BST (SEARCH TITLE)
# =======================
class BSTNode:
    def __init__(self, key):
        self.key = key
        self.song_ids = []
        self.left = None
        self.right = None

class SongBST:
    def __init__(self):
        self.root = None

    def insert(self, key, song_id):
        self.root = self._insert(self.root, key, song_id)

    def _insert(self, node, key, song_id):
        if not node:
            n = BSTNode(key)
            n.song_ids.append(song_id)
            return n
        if key < node.key:
            node.left = self._insert(node.left, key, song_id)
        elif key > node.key:
            node.right = self._insert(node.right, key, song_id)
        else:
            if song_id not in node.song_ids:
                node.song_ids.append(song_id)
        return node

    def search(self, key):
        cur = self.root
        while cur:
            if key < cur.key:
                cur = cur.left
            elif key > cur.key:
                cur = cur.right
            else:
                return cur.song_ids
        return None

# =======================
# LIBRARY
# =======================
class Library:
    def __init__(self):
        self.dll = DoublyLinkedList()
        self.map: Dict[str, DLLNode] = {}
        self.bst = SongBST()
        self.data = []

    def add_song(self, song: Song):
        if song.song_id in self.map:
            return
        node = DLLNode(song)
        self.dll.append(node)
        self.map[song.song_id] = node
        self.bst.insert(song.title.lower(), song.song_id)
        self.data.append(song)

    def get_song(self, sid):
        node = self.map.get(sid)
        return node.data if node else None

    def list_songs(self):
        return self.dll.to_list()

# =======================
# PLAYLIST
# =======================
class Playlist:
    def __init__(self, name):
        self.name = name
        self.dll = DoublyLinkedList()
        self.map = {}

    def add_song_node(self, lib_node: DLLNode):
        sid = lib_node.data.song_id
        if sid in self.map:
            return False
        node = DLLNode(lib_node.data)
        self.dll.append(node)
        self.map[sid] = node
        return True

    def remove_song(self, sid):
        node = self.map.get(sid)
        if not node:
            return False
        self.dll.remove(node)
        del self.map[sid]
        return True

    def list_songs(self):
        return self.dll.to_list()
    
    def get_song_ids(self):
        return [s.song_id for s in self.list_songs()]

# =======================
# PLAYER (PYGAME)
# =======================
pygame.mixer.init()

class Player:
    def __init__(self, library: Library):
        self.library = library
        self.current = None
        self.history = Stack()
        self.playlist_head = None
        self.playlist_mode = False
        self.playlist_ids = None
        self.playlist_index = -1

    def play(self, sid):
        self.playlist_ids = None
        self.playlist_index = -1
        self._play_internal(sid)

        node = self.library.map.get(sid)
        if not node:
            print("Lagu tidak ditemukan")
            return

        path = os.path.join(SONG_DIR, f"{sid}.mp3")
        if not os.path.exists(path):
            print("File lagu tidak ditemukan:", path)
            return

        pygame.mixer.music.load(path)
        pygame.mixer.music.play()

        self.current = node
        print(f"▶ Memutar: {node.data.title} - {node.data.artist}")

    def stop(self):
        pygame.mixer.music.stop()

    def next(self):
        if self.playlist_ids:
            if self.playlist_index + 1 < len(self.playlist_ids):
                self.playlist_index += 1
                self.play(self.playlist_ids[self.playlist_index])
            else:
                print("Sudah lagu terakhir di playlist")
            return
        
        if self.current and self.current.next:
            self.play(self.current.next.data.song_id)

    def prev(self):
        if self.playlist_ids:
            if self.playlist_index > 0:
                self.playlist_index -= 1
                self.play(self.playlist_ids[self.playlist_index])
            else:
                print("Sudah lagu pertama di playlist")
            return
        
        if self.current and self.current.prev:
            self.play(self.current.prev.data.song_id)
    
    def play_from_playlist(self, song_ids):
        self.playlist_mode = True
        self.history = Stack()
        self.playlist_ids = song_ids
        self.playlist_index = 0
        self.play(song_ids[0])
        self._play_internal(song_ids[0])

        prev_node = None
        first_node = None

        for sid in song_ids:
            node = self.library.map.get(sid)
            if not node:
                continue

            node.prev = prev_node
            node.next = None

            if prev_node:
                prev_node.next = node
            else:
                first_node = node

            prev_node = node

        self.playlist_head = first_node
        self.current = first_node

        if self.current:
            self.play(sid)
    
    def _play_internal(self, sid):
        node = self.library.map.get(sid)
        if not node:
            print("Lagu tidak ditemukan")
            return
        
        path = os.path.join(SONG_DIR, f"{sid}.mp3")
        if not os.path.exists(path):
            print("File lagu tidak ditemukan:", path)
            return
        
        pygame.mixer.music.load(path)
        pygame.mixer.music.play()
        self.current = node
        print(f"▶ Memutar: {node.data.title} - {node.data.artist}")
# =======================
# MAIN APP
# =======================
class MusicAppCLI:
    def __init__(self):
        self.library = Library()
        self.playlists: Dict[str, Playlist] = {}
        self.admin_password = "admin123"
        self.load_songs()
        self.load_playlists_from_db()
        self.player = Player(self.library)

    # =======================
    # LOAD DATABASE
    # =======================
    def load_songs(self):
        if not os.path.exists(SONGS_FILE):
            print("songs.xlsx tidak ditemukan!")
            sys.exit()
        wb = load_workbook(SONGS_FILE)
        ws = wb.active
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            sid, title, artist, genre, album, year = row
            self.library.add_song(Song(str(sid), title, artist, genre, album, int(year)))

    def load_playlists_from_db(self):
        plist_file = os.path.join(DB_DIR, "playlists.xlsx")
        plsongs_file = os.path.join(DB_DIR, "playlist_songs.xlsx")
        if not os.path.exists(plist_file) or not os.path.exists(plsongs_file):
            return
        wb = load_workbook(plist_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            self.playlists[name] = Playlist(name)
        wb2 = load_workbook(plsongs_file)
        ws2 = wb2.active
        for row in ws2.iter_rows(min_row=2, values_only=True):
            pname, sid = row
            if pname in self.playlists and sid in self.library.map:
                self.playlists[pname].add_song_node(self.library.map[sid])

    # =======================
    # MAIN MENU
    # =======================
    def run(self):
        while True:
            clear_screen()
            print("=== MENU UTAMA ===")
            print("1. Login Admin")
            print("2. Login User")
            print("0. Keluar")
            c = input("Pilih: ")
            if c == "1":
                self.admin_login()
            elif c == "2":
                self.user_menu()
            elif c == "0":
                self.player.stop()
                sys.exit()

    # =======================
    # ADMIN
    # =======================
    def admin_login(self):
        clear_screen()
        print("=== LOGIN ADMIN ===")
        print("0. Kembali")
        pw = input("Masukkan password admin: ")
        if pw == "0":
            return
        if pw != self.admin_password:
            print("Password salah")
            input("ENTER...")
            return
        self.admin_menu()

    def admin_menu(self):
        while True:
            clear_screen()
            print("--- ADMIN PAGE ---")
            print("1. Lihat semua lagu")
            print("2. Tambah lagu")
            print("3. Ubah data lagu")
            print("4. Hapus lagu")
            print("0. Kembali")
            c = input("Pilih: ")
            if c == "1": self.admin_view_songs()
            elif c == "2": self.admin_add_song()
            elif c == "3": self.admin_edit_song()
            elif c == "4": self.admin_delete_song()
            elif c == "0": return

    def admin_view_songs(self):
        clear_screen()
        for s in self.library.list_songs():
            print(f"{s.song_id} | {s.title} | {s.artist}")
        input("ENTER...")

    def admin_add_song(self):
        clear_screen()
        sid = input("ID Lagu: ")
        title = input("Judul: ")
        artist = input("Artis: ")
        genre = input("Genre: ")
        album = input("Album: ")
        year = int(input("Tahun: "))
        wb = load_workbook(SONGS_FILE)
        ws = wb.active
        ws.append([sid, title, artist, genre, album, year])
        wb.save(SONGS_FILE)
        self.library.add_song(Song(sid, title, artist, genre, album, year))
        input("Lagu ditambahkan. ENTER...")

    def admin_edit_song(self):
        clear_screen()
        sid = input("ID lagu yang diubah: ")
        song = self.library.get_song(sid)
        if not song:
            print("Lagu tidak ditemukan")
            input("ENTER...")
            return
        song.title = input(f"Judul ({song.title}): ") or song.title
        song.artist = input(f"Artis ({song.artist}): ") or song.artist
        wb = load_workbook(SONGS_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == sid:
                row[1].value = song.title
                row[2].value = song.artist
        wb.save(SONGS_FILE)
        input("Data lagu diubah. ENTER...")

    def admin_delete_song(self):
        clear_screen()
        sid = input("ID lagu yang dihapus: ")
        wb = load_workbook(SONGS_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == sid:
                ws.delete_rows(row[0].row)
                break
        wb.save(SONGS_FILE)
        input("Lagu dihapus dari database. ENTER...")

    # =======================
    # USER
    # =======================
    def user_menu(self):
        while True:
            clear_screen()
            print("--- USER PAGE ---")
            print("1. Buat playlist")
            print("2. Tambah lagu ke playlist")
            print("3. Hapus lagu dari playlist")
            print("4. Lihat playlist")
            print("5. Cari")
            print("6. Play lagu (dari library)")
            print("7. Play dari playlist")
            print("8. Hapus playlist")
            print("0. Kembali")
            c = input("Pilih: ")
            if c == "1": self.create_playlist()
            elif c == "2": self.add_song_to_playlist()
            elif c == "3": self.remove_song_from_playlist()
            elif c == "4": self.view_playlists()
            elif c == "5": self.search_song()
            elif c == "6": self.player_control_library()
            elif c == "7": self.play_from_playlist()
            elif c == "8": self.delete_playlist()
            elif c == "0": return

    def search_song(self):
        while True:
            print("\n=== MENU CARI ===")
            print("1. Cari berdasarkan artis")
            print("2. Cari berdasarkan judul")
            print("0. Kembali")
            pilih = input("Pilih: ")

            if pilih == "1":
                key = input("Masukkan nama artis: ").lower()
                hasil = [
                    s for s in self.library.data
                    if key in s.artist.lower()
                ]
                self.tampilkan_hasil_cari(hasil)

            elif pilih == "2":
                key = input("Masukkan judul lagu: ").lower()
                hasil = [
                    s for s in self.library.data
                    if key in s.title.lower()
                ]
                self.tampilkan_hasil_cari(hasil)

            elif pilih == "0":
                return
            else:
                print("Pilihan tidak valid")
    
    def tampilkan_hasil_cari(self, hasil):
        if not hasil:
            print("Tidak ditemukan lagu yang sesuai")
            return
        
        print("\nHasil Pencarian:")
        for s in hasil:
            print(f"{s.song_id} | {s.title} - {s.artist}")

    def player_control_library(self):
        print("\nTekan 0 untuk kembali")
        key = input("Masukkan judul lagu (substring atau judul lengkap): ").lower()

        if key == "0":
            return
        
        hasil =[
            s for s in self.library.data
            if key in s.title.lower()
        ]

        if not hasil:
            print("Lagu tidak ditemukan")
            return
        
        song = hasil[0]

        self.player.play(song.song_id)
        self.player_control()

    def play_from_playlist(self):
        clear_screen()

        if not self.playlists:
            print("Belum ada playlist")
            input("ENTER...")
            return

        print("Daftar playlist:")
        for name in self.playlists:
            print(f"- {name}")

        pname = input("\nNama playlist: ")
        playlist = self.playlists.get(pname)

        if not playlist:
            print("Playlist tidak ditemukan")
            input("ENTER...")
            return

        song_ids = playlist.get_song_ids()
        if not song_ids:
            print("Playlist kosong")
            input("ENTER...")
            return

    # ▶ putar lagu pertama di playlist
        self.player.play_from_playlist(song_ids)

    # 🎮 masuk ke PLAYER CONTROL YANG SAMA
        self.player_control()

    def player_control(self):
        while True:
            if not self.player.current:
                print("Tidak ada lagu yang sedang diputar")
                return

            song = self.player.current.data

            clear_screen()
            print("\n=== PLAYER CONTROL ===")
            print(f"Sekarang: {song.title} – {song.artist} ({song.album}, {song.year})")
            print("1. Stop")
            print("2. Next")
            print("3. Prev")
            print("0. Kembali")

            pilih = input("Pilih: ")

            if pilih == "1":
                self.player.stop()
            elif pilih == "2":
                self.player.next()
            elif pilih == "3":
                self.player.prev()
            elif pilih == "0":
                return
            else:
                print("Pilihan tidak valid")

    def view_playlists(self):
        clear_screen()
        if not self.playlists:
            print("Belum ada playlist")
        for name, pl in self.playlists.items():
            print(f"Playlist: {name}")
            for s in pl.list_songs():
                print(f"  {s.song_id} | {s.title} - {s.artist}")
        input("ENTER...")

    def create_playlist(self):
        name = input("Nama playlist: ")
        if name in self.playlists:
            return
        self.playlists[name] = Playlist(name)
        plist = os.path.join(DB_DIR, "playlists.xlsx")
        if not os.path.exists(plist):
            wb = Workbook(); ws = wb.active; ws.append(["playlist_name"])
        else:
            wb = load_workbook(plist); ws = wb.active
        ws.append([name]); wb.save(plist)

    def add_song_to_playlist(self):
        pname = input("Nama playlist: ")
        sid = input("ID lagu: ")
        pl = self.playlists.get(pname)
        if pl and sid in self.library.map:
            pl.add_song_node(self.library.map[sid])
            pls = os.path.join(DB_DIR, "playlist_songs.xlsx")
            if not os.path.exists(pls):
                wb = Workbook(); ws = wb.active; ws.append(["playlist_name","song_id"])
            else:
                wb = load_workbook(pls); ws = wb.active
            ws.append([pname, sid]); wb.save(pls)

    def remove_song_from_playlist(self):
        pname = input("Nama playlist: ")
        sid = input("ID lagu: ")
        pl = self.playlists.get(pname)
        if pl: pl.remove_song(sid)

    def delete_playlist(self):
        pname = input("Nama playlist: ")
        if pname in self.playlists:
            del self.playlists[pname]

# =======================
# RUN
# =======================
if __name__ == "__main__":
    MusicAppCLI().run()